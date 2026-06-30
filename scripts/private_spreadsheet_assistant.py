#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HOME = Path.home()
PRIVATE_ROOT = HOME / "HermesPrivate"
SPREADSHEET_INBOX = PRIVATE_ROOT / "inbox" / "spreadsheets"
SPREADSHEET_OUTBOX = PRIVATE_ROOT / "outbox" / "spreadsheets"
SEARCH_ROOTS = [HOME / "Desktop", HOME / "Downloads", SPREADSHEET_INBOX]
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}


@dataclass(frozen=True)
class InboundReservationRequest:
    source_path: Path
    product_query: str
    quantity: float
    unit: str
    inbound_date: date


def normalize_text(value: str) -> str:
    return "".join(str(value).strip().lower().split())


def is_spreadsheet_task(text: str) -> bool:
    normalized = normalize_text(text)
    keywords = [
        "桌面",
        "下载",
        "文件",
        "表格",
        "excel",
        "xlsx",
        "易代仓",
        "预约",
        "入库",
        "新增",
        "修改",
        "编辑",
        "回传",
        "发给我",
    ]
    return any(normalize_text(keyword) in normalized for keyword in keywords)


def find_candidate_file(text: str) -> Path:
    normalized = normalize_text(text)
    candidates: list[Path] = []
    for root in SEARCH_ROOTS:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES:
                candidates.append(path)

    scored: list[tuple[int, float, Path]] = []
    for path in candidates:
        name = normalize_text(path.stem)
        score = 0
        if "易代仓" in normalized and "易代仓" in name:
            score += 8
        if "预约" in normalized and "预约" in name:
            score += 6
        for token in re.findall(r"[\w\u4e00-\u9fff]+", text):
            token_norm = normalize_text(token)
            if len(token_norm) >= 2 and token_norm in name:
                score += min(len(token_norm), 5)
        if "桌面" in normalized and path.parent == HOME / "Desktop":
            score += 3
        if "下载" in normalized and path.is_relative_to(HOME / "Downloads"):
            score += 3
        if score > 0:
            scored.append((score, path.stat().st_mtime, path))

    if not scored:
        raise ValueError("没有在桌面、下载目录或 HermesPrivate/inbox/spreadsheets 找到匹配的 Excel 文件。")
    scored.sort(reverse=True)
    return scored[0][2]


def parse_quantity(text: str) -> tuple[float, str]:
    match = re.search(r"(\d+(?:\.\d+)?)\s*(件|箱|袋|个|斤|kg|KG|千克)?", text)
    if not match:
        raise ValueError("没有识别到数量，例如“100件”。")
    quantity = float(match.group(1))
    unit = match.group(2) or "件"
    return quantity, unit


def parse_target_date(text: str, today: date | None = None) -> date:
    today = today or date.today()
    if "明天" in text:
        return today + timedelta(days=1)
    if "后天" in text:
        return today + timedelta(days=2)
    match = re.search(r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})日?", text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    match = re.search(r"(\d{1,2})月(\d{1,2})日?", text)
    if match:
        return date(today.year, int(match.group(1)), int(match.group(2)))
    return today


def parse_product_query(text: str) -> str:
    known_products = ["西兰花", "西蓝花", "菠菜", "土豆", "虾仁", "玉米", "鸡胸", "鸡腿", "牛五花", "三文鱼"]
    normalized = normalize_text(text)
    for product in known_products:
        if normalize_text(product) in normalized:
            return product
    match = re.search(r"入库(.+?)(?:\d+(?:\.\d+)?\s*(?:件|箱|袋|个|斤|kg|KG|千克)?)", text)
    if match:
        return match.group(1).strip(" 的，,。")
    raise ValueError("没有识别到商品名称。")


def parse_inbound_reservation(text: str) -> InboundReservationRequest:
    if not is_spreadsheet_task(text) or "入库" not in text:
        raise ValueError("当前只支持自然语言处理易代仓入库预约表。")
    quantity, unit = parse_quantity(text)
    return InboundReservationRequest(
        source_path=find_candidate_file(text),
        product_query=parse_product_query(text),
        quantity=quantity,
        unit=unit,
        inbound_date=parse_target_date(text),
    )


def load_catalog_from_workbook(workbook: Any) -> list[dict[str, str]]:
    if "Sheet2" not in workbook.sheetnames:
        return []
    sheet = workbook["Sheet2"]
    header_row = None
    headers: dict[str, int] = {}
    for row_index in range(1, min(sheet.max_row, 20) + 1):
        row = [sheet.cell(row_index, col).value for col in range(1, sheet.max_column + 1)]
        header_map = {str(value).strip(): idx for idx, value in enumerate(row) if value not in (None, "")}
        if "项目" in header_map and "编码" in header_map:
            header_row = row_index
            headers = header_map
            break
    if header_row is None:
        return []

    catalog = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        sku = clean_value(value_at(row, headers.get("编码")))
        name = clean_value(value_at(row, headers.get("项目")))
        if not sku or not name:
            continue
        catalog.append(
            {
                "sku": sku,
                "name": name,
                "storage": clean_value(value_at(row, headers.get("存储方式"))),
                "spec": clean_value(value_at(row, headers.get("箱规"))) or clean_value(value_at(row, headers.get("物料规格"))),
                "unit": clean_value(value_at(row, headers.get("单位"))),
            }
        )
    return catalog


def clean_value(value: Any) -> str:
    return "" if value is None else str(value).strip()


def value_at(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def product_aliases(name: str) -> set[str]:
    normalized = normalize_text(name)
    without_brand = re.sub(r"^熊小小牛排饭", "", normalized).lstrip("-－")
    without_mark = re.sub(r"[（(]冻[）)]", "", without_brand)
    return {
        value
        for value in {
            normalized,
            without_brand,
            without_mark,
            without_mark.removesuffix("冻"),
            without_mark.replace("冷冻", "").removesuffix("冻"),
            without_mark.replace("西兰花", "西蓝花"),
            without_mark.replace("西蓝花", "西兰花"),
        }
        if len(value) >= 2
    }


def match_product(catalog: list[dict[str, str]], query: str) -> dict[str, str]:
    normalized_query = normalize_text(query).replace("西蓝花", "西兰花")
    for item in catalog:
        aliases = {alias.replace("西蓝花", "西兰花") for alias in product_aliases(item["name"])}
        if any(normalized_query in alias or alias in normalized_query for alias in aliases):
            return item
    raise ValueError(f"没有在模板商品字典里找到：{query}")


def header_map(sheet: Any) -> dict[str, int]:
    return {
        clean_value(sheet.cell(1, col).value): col
        for col in range(1, sheet.max_column + 1)
        if clean_value(sheet.cell(1, col).value)
    }


def first_empty_data_row(sheet: Any) -> int:
    for row in range(2, max(sheet.max_row, 2) + 1):
        if all(sheet.cell(row, col).value in (None, "") for col in range(1, sheet.max_column + 1)):
            return row
    return sheet.max_row + 1


def copy_row_style(sheet: Any, source_row: int, target_row: int) -> None:
    from copy import copy

    for col in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)


def write_inbound_reservation(request: InboundReservationRequest) -> dict[str, Any]:
    SPREADSHEET_INBOX.mkdir(parents=True, exist_ok=True)
    SPREADSHEET_OUTBOX.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    working_copy = SPREADSHEET_INBOX / f"{request.source_path.stem}_{timestamp}{request.source_path.suffix}"
    shutil.copy2(request.source_path, working_copy)

    workbook = load_workbook(working_copy)
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.worksheets[0]
    headers = header_map(sheet)
    required = ["预计入库日期", "库存地点", "商品编码", "商品名称", "物料规格", "储存方式", "到货数量", "单位"]
    missing = [label for label in required if label not in headers]
    if missing:
        raise ValueError("模板缺少字段：" + "、".join(missing))

    product = match_product(load_catalog_from_workbook(workbook), request.product_query)
    target_row = first_empty_data_row(sheet)
    copy_row_style(sheet, 2 if sheet.max_row >= 2 else 1, target_row)

    values = {
        "预计入库日期": request.inbound_date,
        "库存地点": "成都易代仓",
        "商品编码": product["sku"],
        "商品名称": product["name"],
        "物料规格": product["spec"],
        "储存方式": product["storage"],
        "到货数量": request.quantity,
        "单位": request.unit or product["unit"] or "件",
    }
    for label, value in values.items():
        cell = sheet.cell(target_row, headers[label])
        cell.value = value
        if label == "预计入库日期":
            cell.number_format = "yyyy-mm-dd"

    output_path = SPREADSHEET_OUTBOX / f"{request.source_path.stem}_已处理_{timestamp}.xlsx"
    workbook.save(output_path)

    return {
        "status": "ok",
        "source_path": str(request.source_path),
        "working_copy": str(working_copy),
        "output_path": str(output_path),
        "sheet": sheet.title,
        "row": target_row,
        "date": request.inbound_date.isoformat(),
        "sku": product["sku"],
        "name": product["name"],
        "quantity": request.quantity,
        "unit": request.unit or product["unit"] or "件",
    }


def format_result(result: dict[str, Any]) -> str:
    return "\n".join(
        [
            "表格已处理完成。",
            f"原文件：{result['source_path']}",
            f"处理方式：复制原文件后，在 {result['sheet']} 第 {result['row']} 行新增入库预约。",
            f"新增内容：{result['date']}｜{result['name']}｜{result['quantity']:g}{result['unit']}｜商品编码 {result['sku']}。",
            f"新文件：{result['output_path']}",
            "安全说明：没有覆盖桌面原文件。",
        ]
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Hermes 私人表格助理")
    sub = parser.add_subparsers(dest="command", required=True)
    process = sub.add_parser("process-text", help="按自然语言处理私人表格任务")
    process.add_argument("text", nargs="+")
    process.add_argument("--json", action="store_true")
    args = parser.parse_args()

    if args.command == "process-text":
        request = parse_inbound_reservation(" ".join(args.text))
        result = write_inbound_reservation(request)
        if args.json:
            print(json.dumps(result, ensure_ascii=False, indent=2))
        else:
            print(format_result(result))
        return 0
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
