import importlib.util
import sys
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
PARSER_PATH = ROOT / "inventory-board" / "app" / "parser.py"
spec = importlib.util.spec_from_file_location("inventory_board_parser_for_tests", PARSER_PATH)
parser_module = importlib.util.module_from_spec(spec)
assert spec and spec.loader
sys.modules[spec.name] = parser_module
spec.loader.exec_module(parser_module)
parse_inventory_file = parser_module.parse_inventory_file


def test_inbound_parser_maps_branded_sku_name_to_catalog_code(tmp_path):
    path = tmp_path / "inbound.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["预计入库日期", "库存地点", "供应商编码", "供应商名称", "商品编码", "商品名称", "物料规格", "储存方式", "到货数量", "单位"])
    sheet.append(["2026-06-30", "", "", "", "熊小小牛排饭-冷冻西兰花（冻）", "熊小小牛排饭-冷冻西兰花（冻）", "", "", 3, "件"])

    catalog = workbook.create_sheet("Sheet2")
    catalog.append([None, "熊小小牛排饭"])
    catalog.append(["编码", "项目", "编码", "存储方式", "箱规", "最小包装单元", "单位", "备注"])
    catalog.append(["LDXXX0005", "熊小小牛排饭-冷冻西兰花（冻）", "LDXXX0005", "冷冻", "10kg/箱", "2.5kg*4袋", "件", None])

    workbook.save(path)

    lines = parse_inventory_file(path, "inbound")

    assert len(lines) == 1
    assert lines[0].sku == "LDXXX0005"
    assert lines[0].name == "熊小小牛排饭-冷冻西兰花（冻）"
    assert lines[0].spec == "10kg/箱"
    assert lines[0].unit == "件"
    assert lines[0].warehouse == "冷冻"


def test_inbound_parser_allows_short_product_name_inside_sku(tmp_path):
    path = tmp_path / "inbound-short-name.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["预计入库日期", "库存地点", "供应商编码", "供应商名称", "商品编码", "商品名称", "物料规格", "储存方式", "到货数量", "单位"])
    sheet.append(["2026-06-30", "", "", "", "易代仓-西兰花-批次01", "", "", "", 2, "件"])

    catalog = workbook.create_sheet("Sheet2")
    catalog.append([None, "熊小小牛排饭"])
    catalog.append(["编码", "项目", "编码", "存储方式", "箱规", "最小包装单元", "单位", "备注"])
    catalog.append(["LDXXX0005", "熊小小牛排饭-冷冻西兰花（冻）", "LDXXX0005", "冷冻", "10kg/箱", "2.5kg*4袋", "件", None])

    workbook.save(path)

    lines = parse_inventory_file(path, "inbound")

    assert lines[0].sku == "LDXXX0005"
