#!/usr/bin/env python3
from __future__ import annotations

import importlib.util
from datetime import date
from pathlib import Path
import sys
import tempfile
import unittest

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = ROOT / "scripts" / "private_spreadsheet_assistant.py"


def load_assistant():
    spec = importlib.util.spec_from_file_location("private_spreadsheet_assistant_for_test", SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def build_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["预计入库日期", "库存地点", "供应商编码", "供应商名称", "商品编码", "商品名称", "物料规格", "储存方式", "到货数量", "单位"])
    for _ in range(3):
        sheet.append([None] * 10)
    catalog = workbook.create_sheet("Sheet2")
    catalog.append([None, "熊小小牛排饭"])
    catalog.append(["编码", "项目", "编码", "存储方式", "箱规", "最小包装单元", "单位", "备注"])
    catalog.append(["LDXXX0005", "熊小小牛排饭-冷冻西兰花（冻）", "LDXXX0005", "冷冻", "10kg/箱", "2.5kg*4袋", "件", None])
    workbook.save(path)


class PrivateSpreadsheetAssistantTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.assistant = load_assistant()

    def test_parse_request_from_natural_text(self) -> None:
        quantity, unit = self.assistant.parse_quantity("新增一个明天的入库西兰花 100 件")
        self.assertEqual(quantity, 100)
        self.assertEqual(unit, "件")
        self.assertEqual(
            self.assistant.parse_target_date("明天入库", today=date(2026, 6, 30)).isoformat(),
            "2026-07-01",
        )
        self.assertEqual(self.assistant.parse_product_query("入库西兰花100件"), "西兰花")

    def test_write_inbound_reservation_copies_source_and_adds_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "易代仓预约.xlsx"
            build_template(source)
            self.assistant.SPREADSHEET_INBOX = tmp_path / "inbox"
            self.assistant.SPREADSHEET_OUTBOX = tmp_path / "outbox"

            request = self.assistant.InboundReservationRequest(
                source_path=source,
                product_query="西兰花",
                quantity=100,
                unit="件",
                inbound_date=date(2026, 7, 1),
            )
            result = self.assistant.write_inbound_reservation(request)

            output = Path(result["output_path"])
            self.assertTrue(output.exists())
            original = load_workbook(source)
            edited = load_workbook(output)
            self.assertIsNone(original["Sheet1"]["E2"].value)
            row = [edited["Sheet1"].cell(2, col).value for col in range(1, 11)]
            self.assertEqual(row[1], "成都易代仓")
            self.assertEqual(row[4], "LDXXX0005")
            self.assertEqual(row[5], "熊小小牛排饭-冷冻西兰花（冻）")
            self.assertEqual(row[6], "10kg/箱")
            self.assertEqual(row[7], "冷冻")
            self.assertEqual(row[8], 100)
            self.assertEqual(row[9], "件")


if __name__ == "__main__":
    unittest.main()
