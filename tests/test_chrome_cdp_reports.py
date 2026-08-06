from __future__ import annotations

import importlib.util
import pathlib
import sys
import tempfile
import unittest
import zipfile
from unittest import mock

from openpyxl import Workbook


ROOT = pathlib.Path(__file__).resolve().parents[1]
SCRIPT_PATH = ROOT / "business-report-dashboard" / "chrome_cdp_reports.py"


def load_module():
    spec = importlib.util.spec_from_file_location("chrome_cdp_reports_for_test", SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


class ChromeCdpReportsTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.module = load_module()

    def test_process_reports_without_explicit_inputs_auto_discovers_daily_files(self) -> None:
        with mock.patch.object(self.module.subprocess, "run") as run:
            self.module.process_reports()
        args = run.call_args.args[0]
        self.assertNotIn("--allow-missing-platform", args)

    def test_process_reports_with_one_missing_platform_is_explicitly_partial(self) -> None:
        with mock.patch.object(self.module.subprocess, "run") as run:
            self.module.process_reports(eleme=pathlib.Path("eleme.xlsx"))
        args = run.call_args.args[0]
        self.assertIn("--allow-missing-platform", args)
        self.assertIn("--eleme", args)

    def test_validate_eleme_report_rejects_header_only_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = pathlib.Path(directory) / "eleme.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "data"
            sheet.append(["日期", "门店名称", "订单量"])
            workbook.save(path)

            with self.assertRaisesRegex(self.module.EmptyReportError, "只有表头"):
                self.module.validate_eleme_report_file(path, "20260805")

    def test_validate_eleme_report_accepts_target_date_rows(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = pathlib.Path(directory) / "eleme.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "data"
            sheet.append(["日期", "门店名称", "订单量"])
            sheet.append(["2026-08-05", "安贞店", 12])
            workbook.save(path)

            self.assertEqual(self.module.validate_eleme_report_file(path, "20260805"), 1)

    def test_validate_eleme_report_ignores_stale_worksheet_dimension(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = pathlib.Path(directory) / "eleme.xlsx"
            rewritten = pathlib.Path(directory) / "rewritten.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "data"
            sheet.append(["日期", "门店名称", "订单量"])
            sheet.append(["2026-08-05", "安贞店", 12])
            workbook.save(path)

            with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(rewritten, "w") as target:
                for item in source.infolist():
                    content = source.read(item.filename)
                    if item.filename == "xl/worksheets/sheet1.xml":
                        content = content.replace(b'<dimension ref="A1:C2"/>', b'<dimension ref="A1:A1"/>')
                    target.writestr(item, content)

            self.assertEqual(self.module.validate_eleme_report_file(rewritten, "20260805"), 1)

    def test_wait_for_eleme_report_returns_empty_export_immediately(self) -> None:
        with mock.patch.object(
            self.module,
            "download_eleme_latest",
            side_effect=self.module.EmptyReportError("empty"),
        ) as download:
            with self.assertRaises(self.module.EmptyReportError):
                self.module.wait_for_eleme_report("20260805", timeout_seconds=180)
        download.assert_called_once_with("20260805")


if __name__ == "__main__":
    unittest.main()
