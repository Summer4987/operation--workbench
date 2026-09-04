from __future__ import annotations

import os
import re
import shutil
import json
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from copy import copy


BASE_DIR = Path(__file__).resolve().parents[1]
TEMPLATE_FILENAMES = ("熊小小牛排饭订单模板.xlsx", "熊小小排饭订单模板.xlsx")
LOCAL_TEMPLATE_DIR = Path("/Users/summer/Desktop/库存管理/出入库模板")
PROJECT_TEMPLATE_DIR = BASE_DIR / "data" / "templates"
SERVER_DATA_DIR = Path(os.environ.get("INVENTORY_DATA_DIR", "/opt/inventory-board/data"))
SERVER_TEMPLATE_DIR = SERVER_DATA_DIR / "templates"
LOCAL_OUTPUT_DIR = Path("/Users/summer/Desktop/库存管理/出库记录")
SERVER_OUTPUT_DIR = BASE_DIR / "data" / "order_outputs"
CATALOG_PATH = BASE_DIR / "app" / "catalog.json"
DISABLED_PUBLIC_ORDER_SKUS = {"CWXXX0004", "LDXXX0005"}
DISABLED_PUBLIC_ORDER_NAMES = {"打包袋", "熊小小牛排饭-定制无纺布袋-YDC", "熊小小牛排饭-冷冻西兰花（冻）"}


def _env_path(name: str) -> Path | None:
    raw = os.environ.get(name, "").strip()
    return Path(raw).expanduser() if raw else None


def _template_candidates() -> list[Path]:
    configured = _env_path("INVENTORY_TEMPLATE_PATH")
    candidates = []
    if configured:
        candidates.append(configured)
    candidates.extend(
        [
            PROJECT_TEMPLATE_DIR / TEMPLATE_FILENAMES[0],
            PROJECT_TEMPLATE_DIR / TEMPLATE_FILENAMES[1],
            SERVER_TEMPLATE_DIR / TEMPLATE_FILENAMES[0],
            SERVER_TEMPLATE_DIR / TEMPLATE_FILENAMES[1],
            LOCAL_TEMPLATE_DIR / TEMPLATE_FILENAMES[0],
            LOCAL_TEMPLATE_DIR / TEMPLATE_FILENAMES[1],
        ]
    )
    return candidates


TEMPLATE_PATH = next((path for path in _template_candidates() if path.exists()), PROJECT_TEMPLATE_DIR / TEMPLATE_FILENAMES[0])
OUTPUT_DIR = _env_path("INVENTORY_OUTPUT_DIR") or SERVER_OUTPUT_DIR

FIXED_CUSTOMERS = [
    {
        "name": "金融城店",
        "address": "四川省成都市武侯区石羊街道新街里6c区3楼3035号熊小小牛排饭",
        "contact": "金融城店",
        "phone": "13281037620",
        "aliases": ("金融城店", "金融城", "新街里店", "新街里"),
    },
    {
        "name": "银泰城店",
        "address": "四川省成都市武侯区桂溪街道益州大道1999号成都银泰城悦坊6栋二层222熊小小牛排饭",
        "contact": "王龙辉",
        "phone": "18328316744",
        "aliases": ("银泰城店", "银泰城", "成都银泰城"),
    },
    {
        "name": "万象城店",
        "address": "四川省成都市成华区万年场街道华润柒公馆双福一路58号一层附99号熊小小牛排饭",
        "contact": "冯强",
        "phone": "13547304996",
        "aliases": ("万象城店", "万象城", "柒公馆店", "柒公馆"),
    },
    {
        "name": "保利中心店",
        "address": "四川省成都市武侯区玉林街道保利中心东区C座一层熊小小牛排饭",
        "contact": "陈丹",
        "phone": "18382053718",
        "aliases": ("保利中心店", "保利中心", "保利"),
    },
]

PRODUCT_ALIAS_OVERRIDES = {
    "熊小小牛排饭-冷冻西兰花（冻）": ("西兰花", "西蓝花"),
    "熊小小牛排饭-冷冻菠菜（冻）": ("菠菜",),
    "熊小小牛排饭-牛五花牛排（冻）": ("牛五花", "五花"),
    "熊小小牛排饭-调理鸡胸肉（冻）": ("鸡胸", "鸡胸肉"),
    "熊小小牛排饭-调理手枪腿（冻）": ("手枪腿", "鸡腿"),
    "熊小小牛排饭-三文鱼块（冻）": ("三文鱼",),
}


@dataclass(frozen=True)
class Product:
    sku: str
    name: str
    spec: str
    unit: str
    warehouse: str
    aliases: tuple[str, ...]
    public_order: bool = True


@dataclass(frozen=True)
class Customer:
    name: str
    address: str
    contact: str
    phone: str
    aliases: tuple[str, ...] = ()


@dataclass(frozen=True)
class OrderLine:
    store_name: str
    sku: str
    product_name: str
    spec: str
    unit: str
    warehouse: str
    quantity: Decimal
    address: str
    contact: str
    phone: str
    source_text: str


class OrderParseError(ValueError):
    pass


def load_catalogs(template_path: Path = TEMPLATE_PATH) -> tuple[list[Product], list[Customer]]:
    if not template_path.exists():
        raise OrderParseError(f"没有找到出库模板：{template_path}")

    workbook = load_workbook(template_path, data_only=True)
    products = _merge_products(_load_products(workbook), _load_project_catalog_products())
    customers = _merge_customers(_fixed_customers(), _load_customers(workbook))
    if not products:
        raise OrderParseError("模板里没有读到货品信息")
    if not customers:
        raise OrderParseError("模板里没有读到客户信息")
    return products, customers


def preview_wechat_order(text: str, template_path: Path = TEMPLATE_PATH) -> dict:
    products, customers = load_catalogs(template_path)
    lines, warnings = parse_wechat_order(text, products, customers)
    return {
        "line_count": len(lines),
        "warnings": warnings,
        "items": [
            {
                "store_name": line.store_name,
                "sku": line.sku,
                "product_name": line.product_name,
                "quantity": float(line.quantity),
                "unit": line.unit,
                "address": line.address,
                "source_text": line.source_text,
            }
            for line in lines
        ],
    }


def generate_outbound_order(text: str, template_path: Path = TEMPLATE_PATH, output_dir: Path = OUTPUT_DIR) -> dict:
    products, customers = load_catalogs(template_path)
    lines, warnings = parse_wechat_order(text, products, customers)
    return _generate_outbound_order_from_lines(lines, warnings, template_path, output_dir)


def generate_structured_outbound_order(
    store_name: str,
    items: list[dict],
    template_path: Path = TEMPLATE_PATH,
    output_dir: Path = OUTPUT_DIR,
) -> dict:
    products, customers = load_catalogs(template_path)
    customer = _match_customer(store_name, customers)
    if not customer:
        raise OrderParseError("请选择门店")

    products_by_sku = {product.sku: product for product in products}
    lines: list[OrderLine] = []
    for item in items:
        sku = _clean(item.get("sku"))
        product = products_by_sku.get(sku)
        quantity = _decimal(str(item.get("quantity", "")))
        if not product or not product.public_order or quantity is None or quantity <= 0:
            continue
        lines.append(
            OrderLine(
                store_name=customer.name,
                sku=product.sku,
                product_name=product.name,
                spec=product.spec,
                unit=product.unit,
                warehouse=product.warehouse,
                quantity=quantity,
                address=customer.address,
                contact=customer.contact,
                phone=customer.phone,
                source_text="门店链接提交",
            )
        )

    return _generate_outbound_order_from_lines(lines, [], template_path, output_dir)


def public_order_catalog(template_path: Path = TEMPLATE_PATH) -> dict:
    if template_path.exists():
        products, customers = load_catalogs(template_path)
    else:
        products = _load_project_catalog_products()
        customers = _fixed_customers()
        if not products:
            raise OrderParseError(f"没有找到出库模板：{template_path}")
    return {
        "stores": [
            {
                "name": customer.name,
                "address": customer.address,
                "contact": customer.contact,
                "phone": customer.phone,
            }
            for customer in customers
            if customer.name in {"金融城店", "银泰城店", "万象城店", "保利中心店"}
        ],
        "products": [
            {
                "sku": product.sku,
                "name": product.name.replace("熊小小牛排饭-", ""),
                "spec": product.spec,
                "unit": product.unit,
            }
            for product in products
            if product.public_order and not _is_disabled_public_order_product(product)
        ],
    }


def _generate_outbound_order_from_lines(
    lines: list[OrderLine],
    warnings: list[str],
    template_path: Path,
    output_dir: Path,
) -> dict:
    if not lines:
        raise OrderParseError("没有识别到可生成的订货明细，请检查是否包含门店名、品项和数量。")

    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"熊小小牛排饭订单模板_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = output_dir / filename
    shutil.copyfile(template_path, output_path)

    workbook = load_workbook(output_path)
    if "客户订单填写" not in workbook.sheetnames:
        raise OrderParseError("出库模板里没有找到“客户订单填写”工作表")
    sheet = workbook["客户订单填写"]

    end_row = max(sheet.max_row, len(lines) + 1)
    _clear_input_area(sheet, end_row)
    _ensure_formula_rows(sheet, len(lines) + 1)

    for row_number, line in enumerate(lines, start=2):
        sheet.cell(row_number, 1).value = line.sku
        sheet.cell(row_number, 2).value = line.product_name
        sheet.cell(row_number, 3).value = line.spec
        sheet.cell(row_number, 4).value = line.unit
        sheet.cell(row_number, 5).value = float(line.quantity)
        sheet.cell(row_number, 6).value = line.address
        sheet.cell(row_number, 7).value = line.contact
        sheet.cell(row_number, 8).value = line.phone
        sheet.cell(row_number, 9).value = ""

    workbook.save(output_path)
    return {
        "line_count": len(lines),
        "warnings": warnings,
        "file": str(output_path),
        "items": [
            {
                "store_name": line.store_name,
                "sku": line.sku,
                "product_name": line.product_name,
                "spec": line.spec,
                "quantity": float(line.quantity),
                "unit": line.unit,
                "warehouse": line.warehouse,
                "address": line.address,
            }
            for line in lines
        ],
    }


def parse_wechat_order(text: str, products: list[Product], customers: list[Customer]) -> tuple[list[OrderLine], list[str]]:
    if not text or not text.strip():
        raise OrderParseError("请先粘贴微信订货消息")

    customer_index = {customer.name: customer for customer in customers}
    all_aliases = _product_aliases(products)
    lines: list[OrderLine] = []
    warnings: list[str] = []
    current_customer: Customer | None = None

    for raw_line in text.splitlines():
        source = raw_line.strip()
        if not source:
            continue

        matched_customer = _match_customer(source, customers)
        if matched_customer:
            current_customer = matched_customer
            store_only_text = _remove_customer_name(source, matched_customer).strip()
            if not store_only_text:
                continue
            source_for_items = store_only_text
        else:
            source_for_items = source

        found_any = False
        line_seen_skus: set[str] = set()
        for product, alias in all_aliases:
            if product.sku in line_seen_skus:
                continue
            for quantity in _quantities_near_alias(source_for_items, alias):
                customer = matched_customer or current_customer
                if not customer:
                    warnings.append(f"未识别门店：{source}")
                    found_any = True
                    line_seen_skus.add(product.sku)
                    continue
                lines.append(
                    OrderLine(
                        store_name=customer.name,
                        sku=product.sku,
                        product_name=product.name,
                        spec=product.spec,
                        unit=product.unit,
                        warehouse=product.warehouse,
                        quantity=quantity,
                        address=customer.address,
                        contact=customer.contact,
                        phone=customer.phone,
                        source_text=source,
                    )
                )
                found_any = True
                line_seen_skus.add(product.sku)

        if matched_customer and not found_any and _looks_like_order_line(source):
            warnings.append(f"这行有门店但没识别到商品数量：{source}")
        elif not matched_customer and not found_any and source:
            maybe_name = _guess_unknown_product(source, customer_index)
            if maybe_name:
                warnings.append(f"可能有未匹配商品：{source}")

    deduped = _dedupe_lines(lines)
    return deduped, warnings


def _load_products(workbook) -> list[Product]:
    if "货品信息" not in workbook.sheetnames:
        return []
    sheet = workbook["货品信息"]
    headers = _header_map(sheet[1])
    result = []
    for row in sheet.iter_rows(min_row=2):
        sku = _clean(row[headers.get("货品编号", -1)].value if "货品编号" in headers else "")
        name = _clean(row[headers.get("货品名称", -1)].value if "货品名称" in headers else "")
        if not sku or not name:
            continue
        spec = _clean(row[headers.get("规格", -1)].value if "规格" in headers else "")
        unit = _clean(row[headers.get("单位", -1)].value if "单位" in headers else "")
        warehouse = _clean(row[headers.get("仓库", -1)].value if "仓库" in headers else "")
        result.append(Product(sku=sku, name=name, spec=spec, unit=unit, warehouse=warehouse, aliases=_aliases_for_product(name)))
    return result


def _load_project_catalog_products(path: Path = CATALOG_PATH) -> list[Product]:
    if not path.exists():
        return []
    try:
        rows = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []

    result = []
    for item in rows if isinstance(rows, list) else []:
        sku = _clean(item.get("sku"))
        name = _clean(item.get("name"))
        if not sku or not name:
            continue
        result.append(
            Product(
                sku=sku,
                name=name,
                spec=_clean(item.get("spec")),
                unit=_clean(item.get("unit")) or "件",
                warehouse=_clean(item.get("warehouse")),
                aliases=_aliases_for_product(name),
                public_order=bool(item.get("public_order", True)),
            )
        )
    return result


def _is_disabled_public_order_product(product: Product) -> bool:
    clean_name = product.name.replace("熊小小牛排饭-", "").strip()
    return product.sku in DISABLED_PUBLIC_ORDER_SKUS or product.name in DISABLED_PUBLIC_ORDER_NAMES or clean_name in DISABLED_PUBLIC_ORDER_NAMES


def _merge_products(primary: list[Product], extra: list[Product]) -> list[Product]:
    seen = {product.sku for product in primary}
    merged = list(primary)
    for product in extra:
        if product.sku not in seen:
            merged.append(product)
            seen.add(product.sku)
    return merged


def _load_customers(workbook) -> list[Customer]:
    if "客户信息" not in workbook.sheetnames:
        return []
    sheet = workbook["客户信息"]
    result = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = _clean(row[0] if len(row) > 0 else "")
        address = _clean(row[1] if len(row) > 1 else "") or _clean(row[4] if len(row) > 4 else "")
        contact = _clean(row[2] if len(row) > 2 else "")
        phone = _clean(row[3] if len(row) > 3 else "")
        if name and address:
            result.append(Customer(name=name, address=address, contact=contact, phone=phone, aliases=()))
    return result


def _fixed_customers() -> list[Customer]:
    return [
        Customer(
            name=item["name"],
            address=item["address"],
            contact=item["contact"],
            phone=item["phone"],
            aliases=tuple(item["aliases"]),
        )
        for item in FIXED_CUSTOMERS
    ]


def _merge_customers(primary: list[Customer], secondary: list[Customer]) -> list[Customer]:
    by_name = {customer.name: customer for customer in primary}
    for customer in secondary:
        by_name.setdefault(customer.name, customer)
    return list(by_name.values())


def _header_map(row) -> dict[str, int]:
    return {_clean(cell.value): idx for idx, cell in enumerate(row) if _clean(cell.value)}


def _aliases_for_product(name: str) -> tuple[str, ...]:
    short = re.sub(r"^熊小小牛排饭[-－]?", "", name).strip()
    no_frozen_mark = re.sub(r"[（(]冻[）)]", "", short).strip()
    base = _normalize_name(short)
    aliases = {name, short, no_frozen_mark, base}
    aliases.add(base.removesuffix("冻"))
    aliases.add(base.replace("冷冻", "").removesuffix("冻"))
    aliases.add(base.replace("调理", "").removesuffix("冻"))
    aliases.update(PRODUCT_ALIAS_OVERRIDES.get(name, ()))
    return tuple(sorted({alias for alias in aliases if len(alias) >= 2}, key=len, reverse=True))


def _product_aliases(products: list[Product]) -> list[tuple[Product, str]]:
    aliases: list[tuple[Product, str]] = []
    for product in products:
        aliases.extend((product, alias) for alias in product.aliases)
    return sorted(aliases, key=lambda item: len(item[1]), reverse=True)


def _match_customer(text: str, customers: list[Customer]) -> Customer | None:
    normalized = _normalize_name(text)
    for customer in sorted(customers, key=lambda item: len(item.name), reverse=True):
        names = [customer.name, *customer.aliases]
        for candidate in names:
            name = _normalize_name(candidate)
            if name and (name in normalized or normalized in name):
                return customer

    best: tuple[int, Customer] | None = None
    for customer in customers:
        score = sum(1 for token in _customer_tokens(customer.name) if token and token in normalized)
        if score >= 2 and (best is None or score > best[0]):
            best = (score, customer)
    return best[1] if best else None


def _remove_customer_name(text: str, customer: Customer) -> str:
    result = text
    for name in sorted([customer.name, *customer.aliases], key=len, reverse=True):
        if name and name in result:
            result = result.replace(name, "", 1)
            break
    return result.strip(" ：:，,;-—")


def _customer_tokens(name: str) -> list[str]:
    cleaned = _normalize_name(name)
    return [token for token in re.split(r"(省|市|区|县|店|中心|万达|国芳|商业街)", cleaned) if len(token) >= 2]


def _quantities_near_alias(text: str, alias: str) -> list[Decimal]:
    escaped = re.escape(alias)
    unit = r"(?:件|箱|袋|包|桶|瓶|份|个|斤|kg|KG|千克)?"
    after = re.compile(escaped + r"[\s:：,，xX*×-]*(\d+(?:\.\d+)?)\s*" + unit)
    before = re.compile(r"(\d+(?:\.\d+)?)\s*" + unit + r"[\s:：,，xX*×-]*" + escaped)
    quantities = [_decimal(match.group(1)) for match in after.finditer(text)]
    if not quantities:
        quantities = [_decimal(match.group(1)) for match in before.finditer(text)]
    return [quantity for quantity in quantities if quantity and quantity > 0]


def _decimal(value: str) -> Decimal | None:
    try:
        return Decimal(value)
    except (InvalidOperation, ValueError):
        return None


def _dedupe_lines(lines: list[OrderLine]) -> list[OrderLine]:
    merged: dict[tuple[str, str], OrderLine] = {}
    for line in lines:
        key = (line.store_name, line.sku)
        if key not in merged:
            merged[key] = line
            continue
        old = merged[key]
        merged[key] = OrderLine(
            store_name=old.store_name,
            sku=old.sku,
            product_name=old.product_name,
            spec=old.spec,
            unit=old.unit,
            warehouse=old.warehouse,
            quantity=old.quantity + line.quantity,
            address=old.address,
            contact=old.contact,
            phone=old.phone,
            source_text=f"{old.source_text}；{line.source_text}",
        )
    return list(merged.values())


def _clear_input_area(sheet, end_row: int) -> None:
    for row in range(2, end_row + 1):
        for col in range(1, 10):
            sheet.cell(row, col).value = None


def _ensure_formula_rows(sheet, end_row: int) -> None:
    if end_row <= sheet.max_row:
        return
    for row in range(sheet.max_row + 1, end_row + 1):
        for col in range(1, sheet.max_column + 1):
            source = sheet.cell(2, col)
            target = sheet.cell(row, col)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)
        sheet.cell(row, 10).value = f"=IF(F{row-1}=F{row},J{row-1},J{row-1}+1)" if row > 2 else 1
        sheet.cell(row, 11).value = "=TODAY()"
        sheet.cell(row, 12).value = f"=K{row}&J{row}"
        sheet.cell(row, 13).value = f"=VLOOKUP(A{row},货品信息!A:F,6,0)"
        sheet.cell(row, 14).value = f"=_xlfn.XLOOKUP(F{row},客户信息!B:B,客户信息!A:A)"


def _looks_like_order_line(text: str) -> bool:
    return bool(re.search(r"\d", text))


def _guess_unknown_product(text: str, customer_index: dict[str, Customer]) -> bool:
    if any(name in text for name in customer_index):
        return False
    return bool(re.search(r"[\u4e00-\u9fff]{2,}.*\d", text))


def _normalize_name(value: str) -> str:
    text = _clean(value)
    return re.sub(r"[\s　()（）\-－_]+", "", text)


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()
