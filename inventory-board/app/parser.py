from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


INBOUND_HEADERS = {
    "sku": "商品编码",
    "name": "商品名称",
    "spec": "物料规格",
    "quantity": "到货数量",
    "unit": "单位",
    "warehouse": "库存地点",
    "date": "预计入库日期",
}

OUTBOUND_HEADERS = {
    "sku": "存货编码",
    "name": "名称",
    "spec": "规格",
    "quantity": "数量",
    "unit": "单位",
    "date": "日期",
    "address": "收货地址",
}

STORE_KEYWORDS = [
    ("银泰城", "银泰城店"),
    ("柒公馆", "万象城店"),
    ("新街里", "金融城店"),
    ("保利中心", "保利中心店"),
]


@dataclass(frozen=True)
class ParsedLine:
    movement_type: str
    sku: str
    name: str
    quantity: Decimal
    unit: str = ""
    spec: str = ""
    warehouse: str = ""
    document_date: str = ""
    address: str = ""
    store_name: str = ""
    row_number: int = 0


class ParseError(ValueError):
    pass


def parse_inventory_file(path: Path, movement_type: str) -> list[ParsedLine]:
    if movement_type not in {"inbound", "outbound"}:
        raise ParseError("请选择入库或出库")

    workbook = load_workbook(path, data_only=False)
    if movement_type == "inbound":
        sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.worksheets[0]
        return _parse_sheet(sheet, INBOUND_HEADERS, "inbound")

    sheet_name = "客户订单填写"
    if sheet_name not in workbook.sheetnames:
        raise ParseError("出库模板里没有找到“客户订单填写”这个工作表")
    return _parse_sheet(workbook[sheet_name], OUTBOUND_HEADERS, "outbound")


def parse_product_catalog(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=True)
    if "货品信息" not in workbook.sheetnames:
        return []
    sheet = workbook["货品信息"]
    headers = _header_map(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    required = ["货品编号", "货品名称"]
    if not all(item in headers for item in required):
        return []

    products = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        sku = _clean(row[headers["货品编号"]])
        name = _clean(row[headers["货品名称"]])
        if not sku or not name:
            continue
        products.append(
            {
                "sku": sku,
                "name": name,
                "spec": _clean(_value_at(row, headers.get("规格"))),
                "unit": _clean(_value_at(row, headers.get("单位"))),
                "warehouse": _clean(_value_at(row, headers.get("仓库"))),
            }
        )
    return products


def _parse_sheet(sheet, expected_headers: dict[str, str], movement_type: str) -> list[ParsedLine]:
    merged_values = _merged_values(sheet)
    rows = sheet.iter_rows(values_only=False)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ParseError("Excel 工作表是空的") from exc

    headers = _header_map([cell.value for cell in header_row])
    missing = [label for label in (expected_headers["sku"], expected_headers["quantity"]) if label not in headers]
    if missing:
        raise ParseError(f"模板字段缺失：{', '.join(missing)}")

    lines: list[ParsedLine] = []
    for row in rows:
        values = [_effective_value(cell, merged_values) for cell in row]
        sku = _clean(_value_at(values, headers.get(expected_headers["sku"])))
        quantity_raw = _value_at(values, headers.get(expected_headers["quantity"]))
        quantity = _decimal(quantity_raw)
        if not sku and quantity is None:
            continue
        if not sku:
            raise ParseError(f"第 {row[0].row} 行没有商品编码")
        if quantity is None or quantity <= 0:
            raise ParseError(f"第 {row[0].row} 行数量必须大于 0")

        address = _clean(_value_at(values, headers.get(expected_headers.get("address", ""))))
        lines.append(
            ParsedLine(
                movement_type=movement_type,
                sku=sku,
                name=_clean(_value_at(values, headers.get(expected_headers.get("name", "")))),
                quantity=quantity,
                spec=_clean(_value_at(values, headers.get(expected_headers.get("spec", "")))),
                unit=_clean(_value_at(values, headers.get(expected_headers.get("unit", "")))),
                warehouse=_clean(_value_at(values, headers.get(expected_headers.get("warehouse", "")))),
                document_date=_format_date(_value_at(values, headers.get(expected_headers.get("date", "")))),
                address=address,
                store_name=detect_store_name(address),
                row_number=row[0].row,
            )
        )

    if not lines:
        raise ParseError("没有读到可导入的商品行")
    return lines


def _header_map(values: Iterable) -> dict[str, int]:
    result = {}
    for idx, value in enumerate(values):
        label = _clean(value)
        if label:
            result[label] = idx
    return result


def _merged_values(sheet) -> dict[tuple[int, int], object]:
    values = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        value = sheet.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                values[(row, col)] = value
    return values


def _effective_value(cell, merged_values):
    if cell.value not in (None, ""):
        return cell.value
    return merged_values.get((cell.row, cell.column), cell.value)


def _value_at(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def detect_store_name(address: str) -> str:
    for keyword, store_name in STORE_KEYWORDS:
        if keyword in address:
            return store_name
    return "未识别门店" if address else ""


def _decimal(value) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _format_date(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if text.startswith("="):
        return ""
    return text
